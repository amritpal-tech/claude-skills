"""
Blog Validator — Reusable openpyxl styling helpers.

Import these in your validation script:
    from excel_helpers import fill, font, border, align, section, score_bg, set_col_widths, freeze, col_headers

Colour palette constants are also exported — use them for consistency across tabs.
"""

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colour palette ───────────────────────────────────────────────────────────
C_NAVY    = "1E3A5F"   # primary dark navy — title bars
C_TEAL    = "0F7B8C"   # teal accent — score columns, secondary headers
C_SLATE   = "2D4A6B"   # medium blue-slate — column headers
C_WHITE   = "FFFFFF"
C_OFFWH   = "F8FAFC"   # alternating row background
C_LTGREY  = "E8EDF2"   # max-score cells, neutral fills
C_HEADER  = "2D4A6B"   # default section header bg

# Semantic traffic-light colours
C_PASS    = "166534";  C_LTGREEN  = "DCFCE7"   # green: PASS
C_WARN    = "92400E";  C_LTAMBER  = "FEF9C3"   # amber: WARN
C_FAIL    = "991B1B";  C_LTRED    = "FEE2E2"   # red:   FAIL
C_INFO    = "1E40AF";  C_LTBLUE   = "DBEAFE"   # blue:  informational
C_PURPLE  = "6B21A8";  C_LTPURP   = "F3E8FF"   # purple: E-E-A-T / identity

# Tab colour presets (use as ws.sheet_properties.tabColor)
TAB_NAVY   = C_NAVY
TAB_TEAL   = C_TEAL
TAB_PURPLE = C_PURPLE
TAB_GREEN  = "166534"
TAB_RED    = "DC2626"


# ── Core styling helpers ─────────────────────────────────────────────────────

def fill(hex_color):
    """Solid background fill."""
    return PatternFill("solid", fgColor=hex_color)


def font(bold=False, color="000000", size=11, italic=False):
    """Calibri font with common options."""
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")


def border():
    """Thin light-grey border on all four sides."""
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)


def align(h="left", v="center", wrap=True):
    """Cell alignment with word wrap on by default."""
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ── Compound helpers ─────────────────────────────────────────────────────────

def section(ws, row, label, bg=C_HEADER, ncols=6, color=C_WHITE):
    """
    Render a full-width section header row spanning `ncols` columns.
    The merged cell gets a dark background with white bold text.
    """
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1)
    c.value = label
    c.fill  = fill(bg)
    c.font  = font(True, color, 11)
    c.alignment = align("left")
    c.border    = border()


def col_headers(ws, row, headers, bgs):
    """
    Write a row of column header cells.
    `headers` is a list of strings, `bgs` is a list of hex colour strings.
    Both lists must be the same length.
    """
    for i, (h, bg) in enumerate(zip(headers, bgs), 1):
        c = ws.cell(row, i)
        c.value     = h
        c.fill      = fill(bg)
        c.font      = font(True, C_WHITE, 9)
        c.alignment = align("center")
        c.border    = border()


def score_bg(score, total=5):
    """
    Return (background_hex, font_colour_hex) for a score cell based on
    percentage of total: green ≥78%, amber ≥55%, red below.
    """
    r = score / total
    if r >= 0.78: return C_LTGREEN, C_PASS
    if r >= 0.55: return C_LTAMBER, C_WARN
    return C_LTRED, C_FAIL


def rating_label(score, total=5):
    """Return a short rating string for a score."""
    r = score / total
    if r >= 0.78: return "✅ PASS"
    if r >= 0.55: return "⚠️ WARN"
    return "❌ FAIL"


def set_col_widths(ws, widths):
    """Set explicit column widths from a list (one width per column, left to right)."""
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def freeze(ws, cell="A3"):
    """Freeze panes at `cell` (default: freeze first two rows)."""
    ws.freeze_panes = cell


# ── Data row writer ──────────────────────────────────────────────────────────

def write_criterion_row(ws, row, category, criteria, finding, score, max_score, ncols=6):
    """
    Write a single criterion row across 6 columns:
      col1: category (bold navy)
      col2: criteria text
      col3: finding text (the substantive validation note)
      col4: score (colour-coded)
      col5: max score (grey)
      col6: rating label (colour-coded)

    Alternates row background between white and off-white.
    Row height should be set by the caller (typically 50–55px).
    """
    bg_row = C_OFFWH if row % 2 == 0 else C_WHITE
    sc_bg, sc_fc = score_bg(score, max_score)
    rat = rating_label(score, max_score)

    def _cell(col, value, bold=False, color="000000", h="left", bg=None, sz=9):
        c = ws.cell(row, col)
        c.value     = value
        c.fill      = fill(bg or bg_row)
        c.font      = font(bold, color, sz)
        c.alignment = align(h)
        c.border    = border()

    _cell(1, category,  bold=True, color=C_NAVY, bg=bg_row)
    _cell(2, criteria)
    _cell(3, finding)
    _cell(4, score,     bold=True, color=sc_fc, h="center", bg=sc_bg, sz=10)
    _cell(5, max_score, h="center", bg=C_LTGREY)
    _cell(6, rat,       bold=True, color=sc_fc, h="center", bg=sc_bg)


def write_score_summary_row(ws, row, total, max_total, note="", ncols=6):
    """
    Write a summary row spanning all columns showing total score / max → /5.0.
    """
    pct_score = round((total / max_total) * 5, 2)
    bg, fc = score_bg(pct_score)

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    c1 = ws.cell(row, 1)
    c1.value     = f"TOTAL: {total}/{max_total} points → {pct_score:.1f}/5.0"
    c1.fill      = fill(bg)
    c1.font      = font(True, fc, 12)
    c1.alignment = align("center")
    c1.border    = border()

    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=ncols)
    c2 = ws.cell(row, 4)
    c2.value     = note
    c2.fill      = fill(bg)
    c2.font      = font(False, fc, 9, italic=True)
    c2.alignment = align("left")
    c2.border    = border()

    return pct_score


def tab_score(criteria_list):
    """
    Given a list of (category, criteria, finding, score, max) tuples,
    return the normalised /5.0 tab score.
    """
    total = sum(r[3] for r in criteria_list)
    mx    = sum(r[4] for r in criteria_list)
    return round((total / mx) * 5, 2) if mx else 0


# ── Title bar helper ─────────────────────────────────────────────────────────

def write_title_bar(ws, title, bg=C_NAVY, ncols=6, height=30):
    """Write a full-width title bar in row 1."""
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    t = ws["A1"]
    t.value     = title
    t.fill      = fill(bg)
    t.font      = font(True, C_WHITE, 13)
    t.alignment = align("center")
    t.border    = border()
    ws.row_dimensions[1].height = height
