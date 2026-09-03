#!/usr/bin/env python3
"""
Build the colour-coded validation workbook from findings JSONs.

Usage:
    python3 build_workbook.py --out devcommx/validation/workbooks/DevCommX_Traffic16_Blog_Validation.xlsx \
        devcommx/validation/*.json

Exec Summary tab + one tab per blog. Run this in the MAIN THREAD; subagents cannot
run openpyxl reliably. Scores are recomputed from raw criteria, never taken from an
agent's self-reported verdict.
"""
import os, sys, json, argparse
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook
from excel_helpers import (fill, font, border, align, section, col_headers, score_bg,
                           rating_label, set_col_widths, freeze, write_title_bar,
                           C_NAVY, C_TEAL, C_SLATE, C_WHITE, C_OFFWH, C_LTGREY,
                           C_PASS, C_LTGREEN, C_WARN, C_LTAMBER, C_FAIL, C_LTRED)
from score_findings import score_file, WEIGHTS, verdict

VERDICT_STYLE = {"PUBLISH": (C_PASS, C_LTGREEN),
                 "REVISE":  (C_WARN, C_LTAMBER),
                 "HOLD":    (C_FAIL, C_LTRED)}


def paint_score(cell, score, total=5):
    """score_bg returns (bg_hex, font_hex) - assign both, do not pass the tuple to fill()."""
    bg, fg = score_bg(score, total)
    cell.fill = fill(bg)
    cell.font = font(bold=True, color=fg)
    cell.alignment = align(h="center")
    cell.border = border()
    return cell


def safe_tab(slug, used):
    """Excel tab names: <=31 chars, no []:*?/\\ , unique."""
    name = slug
    for ch in "[]:*?/\\":
        name = name.replace(ch, "-")
    name = name[:31] or "sheet"
    base, n = name, 2
    while name.lower() in used:
        suf = f"~{n}"
        name = base[:31 - len(suf)] + suf
        n += 1
    used.add(name.lower())
    return name


def exec_summary(wb, rows):
    ws = wb.create_sheet("Exec Summary", 0)
    ws.sheet_properties.tabColor = C_NAVY
    write_title_bar(ws, "DevCommX Blog Validation - Executive Summary", ncols=6)

    r = 3
    mean = sum(x["overall"] for x in rows) / len(rows) if rows else 0
    counts = {v: sum(1 for x in rows if x["verdict"] == v) for v in VERDICT_STYLE}
    for label, val in (("Blogs validated", len(rows)),
                       ("Mean weighted score", f"{mean:.2f} / 5.00"),
                       ("Overall verdict", verdict(mean)),
                       ("PUBLISH / REVISE / HOLD",
                        f"{counts['PUBLISH']} / {counts['REVISE']} / {counts['HOLD']}")):
        ws.cell(r, 1, label).font = font(bold=True)
        ws.cell(r, 2, val)
        r += 1

    r += 1
    section(ws, r, "Per-blog scores", ncols=6); r += 1
    col_headers(ws, r, ["Blog", "Words", "Score", "Verdict", "Weakest dimension", "Top fix"],
                [C_SLATE] * 6)
    r += 1
    for x in rows:
        weakest = min(x["dimensions"], key=lambda d: d["mean"]) if x["dimensions"] else None
        ws.cell(r, 1, x["slug"]).alignment = align()
        ws.cell(r, 2, x["words"]).alignment = align(h="center")
        paint_score(ws.cell(r, 3, round(x["overall"], 2)), x["overall"])
        fg, bg = VERDICT_STYLE[x["verdict"]]
        c = ws.cell(r, 4, x["verdict"])
        c.font = font(bold=True, color=fg); c.fill = fill(bg); c.alignment = align(h="center")
        ws.cell(r, 5, f"{weakest['title']} ({weakest['mean']:.1f})" if weakest else "")
        ws.cell(r, 6, x["top_fixes"][0] if x["top_fixes"] else "").alignment = align()
        for col in range(1, 7):
            ws.cell(r, col).border = border()
        r += 1

    # the cross-batch weakness table
    r += 1
    section(ws, r, "Weakest criteria across the set", ncols=6); r += 1
    col_headers(ws, r, ["Criterion", "Mean", "n", "", "", ""], [C_SLATE] * 6); r += 1
    agg = {}
    for x in rows:
        for d in x.get("_raw_dims", []):
            for c in d.get("criteria", []):
                if isinstance(c.get("score"), int):
                    agg.setdefault(c.get("criterion", "?"), []).append(c["score"])
    for m, k, n in sorted((sum(v) / len(v), k, len(v)) for k, v in agg.items())[:8]:
        ws.cell(r, 1, k)
        paint_score(ws.cell(r, 2, round(m, 2)), m)
        ws.cell(r, 3, n).alignment = align(h="center")
        r += 1

    set_col_widths(ws, [52, 9, 9, 12, 30, 70])
    freeze(ws, "A4")


def blog_tab(wb, x, raw, used):
    ws = wb.create_sheet(safe_tab(x["slug"], used))
    ws.sheet_properties.tabColor = VERDICT_STYLE[x["verdict"]][1]
    write_title_bar(ws, x["title"] or x["slug"], ncols=6)

    r = 3
    for label, val in (("Slug", x["slug"]), ("Words", x["words"]),
                       ("Weighted score", f"{x['overall']:.2f} / 5.00"),
                       ("Verdict", x["verdict"])):
        ws.cell(r, 1, label).font = font(bold=True)
        c = ws.cell(r, 2, val)
        if label == "Verdict":
            fg, bg = VERDICT_STYLE[x["verdict"]]
            c.font = font(bold=True, color=fg); c.fill = fill(bg)
        r += 1
    r += 1

    dim_by_key = {d["key"]: d for d in x["dimensions"]}
    for d in raw.get("dimensions", []):
        sc = dim_by_key.get(d.get("key"))
        head = (f"{d.get('title', d.get('key'))}  "
                f"(weight {WEIGHTS.get(d.get('key'), 0):.2f}"
                + (f", {sc['mean']:.1f}/5)" if sc else ")"))
        section(ws, r, head, ncols=6); r += 1
        if d.get("flag"):
            ft = (d.get("flag_type") or "").lower()
            bg = {"pass": C_LTGREEN, "warn": C_LTAMBER, "critical": C_LTRED}.get(ft, C_LTGREY)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            c = ws.cell(r, 1, f"{ft.upper() or 'NOTE'}: {d['flag']}")
            c.fill = fill(bg); c.alignment = align(); c.font = font(italic=True)
            ws.row_dimensions[r].height = 30
            r += 1
        col_headers(ws, r, ["Category", "Criterion", "Finding", "Score", "Max", "Rating"],
                    [C_TEAL] * 6)
        r += 1
        for c_ in d.get("criteria", []):
            s = c_.get("score", 0)
            mx = c_.get("max", 5)
            ws.cell(r, 1, c_.get("category", "")).alignment = align()
            ws.cell(r, 2, c_.get("criterion", "")).alignment = align()
            ws.cell(r, 3, c_.get("finding", "")).alignment = align()
            paint_score(ws.cell(r, 4, s), s, mx)
            ws.cell(r, 5, mx).alignment = align(h="center")
            ws.cell(r, 6, rating_label(s, mx)).alignment = align(h="center")
            for col in range(1, 7):
                ws.cell(r, col).border = border()
            ws.row_dimensions[r].height = 62
            r += 1
        r += 1

    if x["top_fixes"]:
        section(ws, r, "Top fixes", ncols=6); r += 1
        for i, f in enumerate(x["top_fixes"], 1):
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            c = ws.cell(r, 1, f"{i}. {f}")
            c.alignment = align(); c.fill = fill(C_OFFWH if i % 2 else C_WHITE)
            ws.row_dimensions[r].height = 30
            r += 1

    set_col_widths(ws, [16, 30, 95, 8, 7, 12])
    freeze(ws, "A4")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("files", nargs="+")
    ap.add_argument("--out", required=True)
    a = ap.parse_args()

    rows, raws = [], {}
    for p in a.files:
        x = score_file(p)
        raw = json.load(open(p, encoding="utf-8"))
        x["_raw_dims"] = raw.get("dimensions", [])
        rows.append(x)
        raws[x["slug"]] = raw
        for prob in x["problems"]:
            print(f"  !! {x['slug']}: {prob}", file=sys.stderr)
    rows.sort(key=lambda r: -r["overall"])

    wb = Workbook()
    wb.remove(wb.active)
    exec_summary(wb, rows)
    used = {"exec summary"}
    for x in rows:
        blog_tab(wb, x, raws[x["slug"]], used)

    os.makedirs(os.path.dirname(os.path.abspath(a.out)), exist_ok=True)
    wb.save(a.out)
    mean = sum(r["overall"] for r in rows) / len(rows) if rows else 0
    print(f"Wrote {a.out} - {len(rows)} blogs, mean {mean:.2f}, {len(wb.sheetnames)} tabs")


if __name__ == "__main__":
    main()
